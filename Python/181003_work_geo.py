
#
# Geo Infomation value Tracking your Area
# ver. 1.0.0
# subjt. Geo Based Find Area
# langs. python
# maker. ray na
# group. exciting
# dates. 2018. 09. 13.
#

import openpyxl
from selenium import webdriver
from bs4 import BeautifulSoup


# 1. input path to your web driver
driver = webdriver.Chrome('/Documents/chromedriver')
driver.implicitly_wait(1)

# 2-1. Capture to your Area using Search Engine
wb2 = openpyxl.load_workbook("./geoCodingDetail.xlsx")
ws2 = wb2.active

# 2-2. Shot ! Shot ! Shot !

cnt = 0

for x in ws2.rows :
    cnt = int(x[0].row)

for x in range(0, cnt) :
    zone_Name = ws2.cell(row = x+1, column = 1).value

    driver.get('http://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=1&ie=utf8&query=' + str(zone_Name) )
    driver.save_screenshot("area" + str(int(x+1)) )


# 3-1. Geo info Based to find your Area
wb = openpyxl.load_workbook("./geoCodingDetail.xlsx")
ws = wb.active


# 3-2. I Tracking You !

# cnt = 0
#
# for x in ws.rows :
#     cnt = int(x[0].row)
#
# for x in range(0, cnt) :
#     #print(x)
#     lonX = ws.cell(row = x+1, column = 1).value
#     latY = ws.cell(row = x+1, column = 2).value
#
#     print(lonX, latY)
#
#     # url
#     driver.get('http://map.esran.com/')
#
#     driver.find_element_by_name('locname').send_keys(  str(lonX), " ", str(latY) )
#     driver.find_element_by_name('find').click()
#
#     element = driver.find_element_by_css_selector('div.gm-style-iw div div')
#     name = element.text
#     print(name)
